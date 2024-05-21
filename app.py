import openpyxl
from openpyxl.styles import Font
import os

os.chdir(r"\Users\Foveo\Desktop\精舍資料\DATA")
wb = openpyxl.load_workbook('112-2.xlsx')
sheet = wb.worksheets[0]

# price_updates_dict = {'Garlic': 1.99}

print("Processing...")
for rowNum in range(2, sheet.max_row, 1):
    produceName = sheet.cell(rowNum, 1).value
    if produceName is not None:
        print(produceName)
    # if produceName in price_updates_dict:
    #     sheet.cell(rowNum, 2).value = price_updates_dict[produceName]
    #     sheet.cell(rowNum, 2).font = Font(color='FF0000')
# 將結果另存新檔
# wb.save('produceSales_update.xlsx')
print("Done!")