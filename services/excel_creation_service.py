import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pz.models.excel_creation_model import ExcelCreationModelInterface


class ExcelCreationService:
    wb: Workbook
    sheet: Worksheet
    headers: list[str]

    def __init__(self, model: ExcelCreationModelInterface):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

        self.headers = model.get_excel_headers()

        for i, header in enumerate(self.headers):
            self.sheet.cell(row=1, column=i + 1).value = header

    def __del__(self):
        self.wb.close()

    def save(self, filename: str):
        if filename.endswith('.xlsx'):
            self.wb.save(filename)
        else:
            self.wb.save(f'{filename}.xlsx')

    def write_data(self, suppliers):
        row_number = 1

        for supplier in suppliers:
            row_number += 1
            datum = supplier()

            for i, _ in enumerate(self.headers):
                self.sheet.cell(row=row_number, column=i + 1).value = datum[i]
