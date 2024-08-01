import os

import openpyxl

from pz.characterize import CharacterizeCodec


class CharacterizeService:
    def __init__(self):
        pass

    @staticmethod
    def processing_file(excel_file: str, decode: bool, home_phone: bool, hyphen: bool) -> str:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        header_row = 1
        headers: dict[str, int] = {}

        for column in range(1, sheet.max_column + 1, 1):
            cell = sheet.cell(header_row, column)
            headers[cell.value] = column

        phone_key = '住家電話' if home_phone else '行動電話'
        store_key = '法名'

        if phone_key in headers and store_key in headers:
            phone_index = headers[phone_key]
            store_index = headers[store_key]

            for row in range(header_row + 1, sheet.max_row + 1, 1):
                if decode:
                    encoded = sheet.cell(row, store_index).value
                    phone_number = ''
                    if encoded is not None and len(encoded) == 2:
                        phone_number = CharacterizeCodec.decoder(encoded, hyphen=hyphen)

                    if phone_number != '' and phone_number.startswith('0'):
                        sheet.cell(row, phone_index).value = phone_number
                    else:
                        sheet.cell(row, phone_index).value = ''
                else:
                    phone = sheet.cell(row, phone_index).value
                    pure_number = None
                    if phone is not None and phone != '':
                        phone = ''.join(char for char in str(phone) if char.isdigit())
                        if phone.startswith('0') and 9 <= len(phone) <= 10:
                            pure_number = phone

                    if pure_number is not None:
                        encoded = CharacterizeCodec.encoder(pure_number)
                        back = CharacterizeCodec.decoder(encoded)
                        if back == pure_number:
                            print(encoded, phone)
                            sheet.cell(row, store_index).value = encoded
                        else:
                            raise Exception(f'{phone} failed to decode')
                    else:
                        sheet.cell(row, store_index).value = ''

        saved_file = f'{os.path.dirname(excel_file)}/{'decode' if decode else 'encode'}-{os.path.basename(excel_file)}'

        wb.save(saved_file)
        return saved_file
