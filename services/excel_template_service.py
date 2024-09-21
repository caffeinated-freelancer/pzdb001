import os
from typing import Callable, Any

from loguru import logger
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter

from pz.config import PzProjectExcelSpreadsheetConfig
from pz.models.excel_model import ExcelModelInterface
from pz.models.text_with_properties import TextWithProperties
from pz.utils import tuple_to_coordinate
from services.excel_workbook_service import ExcelWorkbookService, PzCellProperties


class ExcelTemplateService:
    service: ExcelWorkbookService
    template: PzProjectExcelSpreadsheetConfig
    target_folder: str
    template_file: str
    destination_path: str
    template_sheet_name: str
    headers: dict[str | int, int]
    header_row: int
    data_skip_row: int
    insert_row_after: int
    page_mode: bool

    def __init__(self, model: ExcelModelInterface, template_spreadsheet: PzProjectExcelSpreadsheetConfig,
                 source_file_name: str, target_folder: str, output_filename_prefix: str,
                 debug: bool = False):
        if not os.path.exists(target_folder):
            raise FileNotFoundError(target_folder)

        self.template = template_spreadsheet
        self.template_file = template_spreadsheet.spreadsheet_file
        self.target_folder = target_folder
        self.data_skip_row = template_spreadsheet.data_skip_row
        self.insert_row_after = template_spreadsheet.insert_row_after

        # Define the source and destination paths
        self.destination_path = f'{target_folder}/{output_filename_prefix}-{os.path.basename(source_file_name)}'

        # Copy the file
        # shutil.copyfile(template_file, destination_path)

        self.service = ExcelWorkbookService(model, self.template_file, header_at=template_spreadsheet.header_row,
                                            header_on_blank_try=self.template.header_on_blank_try,
                                            page_mode=self.template.page_mode,
                                            debug=debug)

        self.page_mode = self.template.page_mode
        self.headers = self.service.get_headers()
        self.header_row = self.service.get_header_row()
        logger.debug(f'Headers: {self.headers}')

    def get_headers(self) -> dict[str | int, int]:
        return self.headers

    def rehash_header(self):
        self.service.rehash_header(header_on_blank_try=self.template.header_on_blank_try)

        self.headers = self.service.get_headers()
        self.header_row = self.service.get_header_row()

    def write_cell_at(self, row: int, column: int, value: str, font=None):
        self.service.write_cell(row, column, value, font=font)

    def get_sheet(self):
        return self.service.sheet

    def add_page_break(self, row_num: int):
        self.service.add_page_break(row_num)

    def write_cell(self, row: int, data: dict[str, str | int | None]):
        for column_name, index in self.headers.items():
            if column_name in data:
                self.write_cell_at(row, index, data[column_name])

    def insert_columns(self, template_index, index: int, amount: int):

        self.service.sheet.insert_cols(idx=index, amount=amount)

        # print(f'insert columns at {index} -> {index + amount - 1}')

        saved_merged_cells: list[tuple[int, int, int, int]] = []

        for merged_range in self.service.sheet.merged_cells:
            min_col, min_row, max_col, max_row = merged_range.bounds

            if index <= min_col:
                saved_merged_cells.append((min_col, min_row, max_col, max_row))

        if len(saved_merged_cells) > 0:
            for merged_cell in saved_merged_cells:
                min_col, min_row, max_col, max_row = merged_cell
                cell1 = self.service.get_cell(min_row, min_col)
                cell2 = self.service.get_cell(max_row, max_col)
                self.service.sheet.unmerge_cells(f'{cell1.coordinate}:{cell2.coordinate}')

        # template_cell = self.service.get_cell(1, template_index)
        column_dimension = self.service.get_column_dimension(get_column_letter(template_index))

        for column in range(template_index, index + amount, 1):
            self.service.set_column_width(column, column_dimension.width)

        # for row in range(1, self.header_row + self.data_skip_row + 3):
        #     print(f'[Row: {row}] for column in range({self.service.max_column + amount - 1} down to {index + 1})')
        #
        #     for column in range(self.service.max_column + amount - 1, index + amount - 1, - 1):
        #         # print(f'copy from row: {row}, column: {column - amount} to {column} ({index})')
        #         source_cell = self.service.get_cell(row, column - amount)
        #         target_cell = self.service.get_cell(row, column)
        #         # self.service.set_cell_properties(row, column, source_cell)
        #
        #         try:
        #             print(f'copy from {source_cell.column_letter} to {target_cell.column_letter}')
        #             if row == 1:
        #                 column_dimension = self.service.get_column_dimension(source_cell.column_letter)
        #                 self.service.set_column_dimension(target_cell.column_letter, column_dimension)
        #         except AttributeError:
        #             pass

        if len(saved_merged_cells) > 0:
            for merged_cell in saved_merged_cells:
                min_col, min_row, max_col, max_row = merged_cell
                cell1 = self.service.get_cell(min_row, min_col + amount)
                cell2 = self.service.get_cell(max_row, max_col + amount)
                self.service.sheet.merge_cells(f'{cell1.coordinate}:{cell2.coordinate}')

        for row in range(1, self.header_row + self.data_skip_row + 3):
            template_cell = self.service.get_cell(row, template_index)

            for column in range(index, index + amount, 1):
                self.service.set_cell_properties(row, column, template_cell)

    def max_page_entries(self):
        return self.service.page_max_row - self.service.header_row

    def write_data(self, suppliers,
                   caller: Any | None = None,
                   callback: Callable[[dict[str | int, str | int], Any, Any], tuple[Any, bool]] | None = None,
                   duplicate_callback: Callable[[dict[str | int, str | int], Any, Any], bool] | None = None,
                   on_page: int = 0,
                   save: bool = True):
        sample_row_num = self.header_row + 1 + self.data_skip_row
        row_num = sample_row_num + (self.service.sheet_max_row - self.service.header_row) * on_page
        counter = 0

        # fonts = [self.service.get_font(row_num, i) for i in range(1, len(self.headers), 1)]
        template_cells = [PzCellProperties(self.service.get_cell(sample_row_num + 1, i)) for i in
                          range(1, len(self.headers) + 1, 1)]

        header_template_cells = [PzCellProperties(self.service.get_cell(self.header_row, i)) for i in
                                 range(1, len(self.headers) + 1, 1)]

        dimension = self.service.get_row_dimensions(row_num)

        callback_data_holder = None

        data_sn = 0

        for supplier in suppliers:
            datum = supplier()

            data_sn += 1

            if self.page_mode and data_sn > self.max_page_entries():
                logger.error(f'資料超過頁面大小')
                break


            if 0 < self.insert_row_after <= counter:
                self.service.sheet.insert_rows(idx=row_num, amount=1)

            if callback is not None:
                callback_data_holder, add_page_break = callback(datum, callback_data_holder, caller)

                if add_page_break:
                    self.add_page_break(row_num - 1)

            if duplicate_callback is not None:
                if duplicate_callback(datum, callback_data_holder, caller):
                    for column_name, index in self.headers.items():
                        self.service.set_cell_properties_from_pz_cell(row_num, index, header_template_cells[index - 1])
                        header_cell_value = self.service.get_cell(self.header_row, index)
                        self.service.write_cell(row_num, index, header_cell_value.value)
                    self.service.set_row_dimensions(row_num, self.service.get_row_dimensions(self.header_row))
                    row_num += 1
                callback_data_holder = datum

            # logger.info(self.headers.items())

            for column_name, index in self.headers.items():
                if index <= len(template_cells):
                    self.service.set_cell_properties_from_pz_cell(row_num, index, template_cells[index - 1])

                try:
                    if column_name in datum:
                        value = datum[column_name]

                        try:
                            if isinstance(value, TextWithProperties):
                                if 'color' in value.properties:
                                    self.service.write_cell(row_num, index, value.text, color=value.properties['color'])
                                else:
                                    self.service.write_cell(row_num, index, value.text)
                            else:
                                self.service.write_cell(row_num, index, value)
                        except AttributeError as e:
                            logger.warning(e)
                        # print(row_num, self.data_skip_row)
                        # print(index, len(template_cells))
                    else:
                        self.service.write_cell(row_num, index, '')
                except AttributeError as e:
                    if self.page_mode:
                        logger.warning(f'{row_num} vs {data_sn} vs {self.service.page_max_row} - {self.service.header_row}')
                    raise e
            if not self.page_mode:
                self.service.set_row_dimensions(row_num, dimension)
            counter += 1
            # logger.info(f'row: {row_num}: {counter} {self.service.page_max_row}')
            row_num += 1

        if not self.page_mode:
            for r in range(row_num, self.service.max_row() + 1):
                for column_name, index in self.headers.items():
                    self.service.write_cell(r, index, '')
                    try:
                        self.service.set_cell_properties_from_pz_cell(row_num, index, template_cells[index - 1])
                    except IndexError as e:
                        logger.warning(f"row: {row_num}, index: {index}, template size: {len(template_cells)}: {e}")
                self.service.set_row_dimensions(row_num, dimension)

        if save:
            self.service.save_as(self.destination_path)

    def save_as(self, file_name: str | None = None):
        if file_name is not None:
            self.service.save_as(file_name)
        else:
            self.service.save_as(self.destination_path)

    def sheet_rename(self, new_name: str):
        self.service.sheet_rename(new_name)

    def duplicate_sheet(self, source: str, target: str):
        self.service.duplicate_sheet(source, target)
        # new_sheet = workbook.copy_worksheet(sheet_to_duplicate)

    def set_template_sheet(self, template_sheet_name):
        self.template_sheet_name = template_sheet_name

    def duplicate_page(self, page_no: int):
        header_row = self.service.header_row
        max_row = self.service.sheet_max_row
        page_max_row = self.service.page_max_row

        row_num = (max_row - header_row) * page_no + header_row + 1

        self.add_page_break((max_row - header_row) * page_no + header_row)

        i = 0
        for row in range(header_row + 1, max_row + 1):
            template_cells = [PzCellProperties(self.service.get_cell(row, x)) for x in
                              range(1, len(self.headers) + 1, 1)]
            dimension = self.service.get_row_dimensions(row)

            # logger.info(f"duplicate from row: {row} to {row_num + i}")

            merged_cell_from = 0

            for x in range(1, len(self.headers) + 1, 1):
                if x <= len(template_cells):
                    self.service.set_cell_properties_from_pz_cell(row_num + i, x, template_cells[x - 1])
                    if i >= page_max_row - header_row:
                        cell = self.service.get_cell_from_default(row, x)
                        value = cell.value
                        if isinstance(cell, MergedCell):
                            logger.trace(f"({row_num},{x}) MergedCell ({row},{x})")
                            if merged_cell_from == 0:
                                merged_cell_from = x - 1
                        else:
                            if merged_cell_from != 0:
                                f = tuple_to_coordinate(row_num + i, merged_cell_from)
                                t = tuple_to_coordinate(row_num + i, x - 1)
                                self.service.sheet.merge_cells(f'{f}:{t}')
                                merged_cell_from = 0

                            if value is not None:
                                logger.trace(f"({row_num},{x}) with {value} ({row},{x})")
                                self.service.get_cell(row_num + i, x).value = value

            self.service.set_row_dimensions(row_num + i, dimension)
            i += 1

    def set_active_sheet(self, sheet_name: str):
        self.service.wb.active = self.service.wb[sheet_name]

    def remove_sheet(self, sheet_name: str):
        self.service.wb.remove(self.service.wb[sheet_name])
