import os.path
import re
from collections import OrderedDict
from copy import copy
from typing import Any, Callable

import openpyxl
from loguru import logger
from openpyxl import Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import Font
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from pz.models.excel_model import ExcelModelInterface
from pz.utils import format_phone_number


class PzCellProperties:
    def __init__(self, another_cell: Cell):
        self.alignment = copy(another_cell.alignment)
        self.font = copy(another_cell.font)
        self.border = copy(another_cell.border)
        self.fill = copy(another_cell.fill)


class ExcelWorkbookService:
    headers: OrderedDict[str, int]
    model: ExcelModelInterface
    headers_rev: dict[Any, Any]
    max_column: int
    sheet: Worksheet | Any
    default_sheet: Worksheet | Any
    debug: bool
    wb: Workbook | None
    header_row: int
    ignore_parenthesis: bool
    print_header: bool
    excel_file_name: str
    default_sheet_title: str
    page_max_row: int
    sheet_max_row: int

    def __init__(self, model: ExcelModelInterface, excel_file: str, sheet_name: str | None = None,
                 header_at: int | None = None, header_on_blank_try: int = 0,
                 page_mode: bool = False,
                 ignore_parenthesis: bool = False, debug: bool = False,
                 print_header: bool = False, read_only: bool = False) -> None:
        self.ignore_parenthesis = ignore_parenthesis
        self.print_header = print_header

        self.wb = openpyxl.load_workbook(excel_file, read_only=read_only)

        self.excel_file_name = os.path.basename(excel_file)

        self.debug = debug
        self.model: ExcelModelInterface = model

        # if debug:
        #     print(self.wb.worksheets)
        #     print(self.wb.sheetnames)
        self.sheet = self.wb[sheet_name] if sheet_name is not None else self.wb.active

        logger.trace(f'Sheet Name: {self.sheet.title}')

        self.default_sheet = self.sheet
        self.default_sheet_title = self.sheet.title

        # self.sheet = self.wb.worksheets[sheet_name]

        self.header_row = 1

        if not read_only and self.sheet.freeze_panes is not None:
            # print(self.sheet.freeze_panes)
            frozen_row, frozen_col = coordinate_to_tuple(self.sheet.freeze_panes)
            # print(frozen_row, frozen_col)
            self.header_row = frozen_row - 1

        if header_at is not None:
            self.header_row = header_at

        # for column in range(1, self.sheet.max_column):
        #     letter = get_column_letter(column)
        #
        #     print(f'{letter}: {self.sheet.column_dimensions[letter].width}')

        logger.debug(f'{excel_file}: {self.wb.sheetnames}/row:{self.header_row} - {self.wb.worksheets}')

        self.rehash_header(header_on_blank_try=header_on_blank_try)

        if page_mode:
            self.page_max_row = self.calculate_page()
        else:
            self.page_max_row = 0

        self.sheet_max_row = self.sheet.max_row

        logger.info(
            f'header_row: {self.header_row}, page_max_row: {self.page_max_row}, sheet_max_row: {self.sheet_max_row}')

    def __del__(self):
        self.close()

    def close(self):
        if self.wb:
            self.wb.close()
            logger.debug(f'關閉 Excel 檔案: {self.excel_file_name}')
            # try:
            #     self.wb = openpyxl.load_workbook('null')
            # finally:
            #     pass
            # self.wb = openpyxl.Workbook()
            # self.wb.close()
            self.wb = None

    def rehash_header(self, header_on_blank_try: int = 0):
        self.headers = OrderedDict()

        self.headers_rev = {}
        self.max_column = 0

        for colNum in range(1, self.sheet.max_column + 1, 1):
            cell = self.sheet.cell(self.header_row, colNum)
            value = cell.value

            if value is None and header_on_blank_try > 0:
                cell = self.sheet.cell(header_on_blank_try, colNum)
                value = cell.value
                # logger.warning(f'{colNum}: {value}')

            if value is not None:
                # logger.warning(f'{colNum}: {value} ({header_on_blank_try})')
                if isinstance(value, str):
                    value = cell.value.replace('\r', '').replace('\n', '')
                    value = value.strip()
                # print("[", value, "]")
                if self.ignore_parenthesis:
                    matched = re.match(r'^\s*(.*\S)\s*\(.*', value)
                    if matched:
                        value = matched.group(1)
                self.headers[value] = colNum
                if isinstance(value, str):
                    self.headers_rev[colNum] = value.strip()
                else:
                    self.headers_rev[colNum] = value
                self.max_column = max(self.max_column, colNum)
                # print(value, max_column)

        self.max_column += 1

        if self.print_header:
            print(f'Header row at: {self.header_row}')
            for header in self.headers:
                # print(header, self.headers[header])
                print(f'\'\': \'{header}\',')

    def calculate_page(self) -> int:
        for rowNum in range(self.header_row + 1, self.sheet.max_row + 1):
            for colNum in range(2, 5):
                cell = self.sheet.cell(row=rowNum, column=colNum)
                if isinstance(cell, MergedCell):
                    logger.debug(f'Merged cell found at {rowNum},{colNum}')
                    return rowNum - 1
        return 0

    def read_row_by_row(self, callback: Callable[[int, list[Cell]], Any]):
        for rowNum in range(self.header_row + 1, self.sheet.max_row + 1):
            cells: list[Cell] = []
            for colNum in range(1, self.sheet.max_column + 1):
                cell = self.sheet.cell(row=rowNum, column=colNum)
                cells.append(cell)
            callback(rowNum, cells)

    def add_page_break(self, row: int):
        self.sheet.row_breaks.append(Break(id=row))

    def read_all(self, required_attribute: str | None = None) -> list[ExcelModelInterface]:
        counter = 0
        blank_counter = 0
        if self.debug:
            logger.debug(self.headers)
        infos = []
        for rowNum in range(self.header_row + 1, self.sheet.max_row + 1, 1):
            counter += 1
            item = {}
            for colNum in range(1, self.max_column, 1):
                if colNum in self.headers_rev:
                    cell = self.sheet.cell(rowNum, colNum)
                    value = cell.value
                    if isinstance(value, str):
                        item[self.headers_rev[colNum]] = value.strip()
                    elif isinstance(value, int):
                        item[self.headers_rev[colNum]] = value

                        if cell.has_style:
                            number_format = cell.number_format
                            if number_format is not None:
                                if number_format == r'0000\-000000':
                                    item[self.headers_rev[colNum]] = format_phone_number(value)
                                elif number_format != 'General':
                                    logger.warning(f'[第 {rowNum} 列, 第 {colNum} 行] 本程式不支援此儲存格格式: {number_format}')
                                    item[self.headers_rev[colNum]] = str(value)
                    else:
                        item[self.headers_rev[colNum]] = value

                # print(cell.value, end='  ')
            # info = PzContactInfo(item)
            info = self.model.new_instance(item)

            if required_attribute is not None:
                if hasattr(info, required_attribute) and info.__dict__.get(required_attribute) is not None:
                    blank_counter = 0
                    # print(info.to_json())
                    infos.append(info)
                else:
                    blank_counter += 1

                    if blank_counter > 10:
                        break
            else:
                infos.append(info)

        # if produceName in price_updates_dict:
        #     sheet.cell(rowNum, 2).value = price_updates_dict[produceName]
        #     sheet.cell(rowNum, 2).font = Font(color='FF0000')
        # 將結果另存新檔
        # wb.save('produceSales_update.xlsx')
        logger.debug(f'records: {counter}, effected records: {len(infos)}, sheet_max_row: {self.max_row()}')
        return infos

    def write_cell(self, row_num, col_num, value, color: str | None = None, font=None):
        cell = self.sheet.cell(row_num, col_num)

        if isinstance(cell, MergedCell):
            logger.warning(f'Merged cell({row_num},{col_num}): {cell}')

        if color is not None:
            current_font = cell.font
            # print(f'set color as {color}')
            # font.setColor(color)
            cell.font = Font(color=color, family=current_font.family,
                             size=current_font.size,
                             bold=current_font.bold, italic=current_font.italic,
                             underline=current_font.underline,
                             vertAlign=current_font.vertAlign,
                             outline=current_font.outline,
                             shadow=current_font.shadow, strike=current_font.strike,
                             charset=current_font.charset, condense=current_font.condense)
        elif font is not None:
            cell.font = font

        try:
            cell.value = value
        except AttributeError as e:
            logger.error(f'write_cell({row_num},{col_num}) = {value}: {e}')
            raise e


    def get_font(self, row_num, col_num) -> Font:
        # col_num = col_num if col_num > 0 else 1
        font = self.sheet.cell(row_num, col_num).font

        # print(f'row:{row_num}, col:{col_num}, size:{font.size}')

        return Font(family=font.family, size=font.size,
                    bold=font.bold, italic=font.italic, underline=font.underline,
                    vertAlign=font.vertAlign, outline=font.outline,
                    shadow=font.shadow, strike=font.strike,
                    charset=font.charset, condense=font.condense)

    def get_cell(self, row_num, col_num) -> Cell | MergedCell:
        # col_num = col_num if col_num > 0 else 1
        return self.sheet.cell(row_num, col_num)

    def get_cell_from_default(self, row_num, col_num) -> Cell | MergedCell:
        # col_num = col_num if col_num > 0 else 1
        return self.default_sheet.cell(row_num, col_num)

    def set_cell_properties(self, row_num, col_num, another_cell: Cell):
        cell = self.sheet.cell(row_num, col_num)

        cell.alignment = copy(another_cell.alignment)
        cell.font = copy(another_cell.font)
        cell.border = copy(another_cell.border)
        cell.fill = copy(another_cell.fill)

    def set_cell_properties_from_pz_cell(self, row_num, col_num, pz_cell: PzCellProperties):
        cell = self.sheet.cell(row_num, col_num)

        cell.alignment = copy(pz_cell.alignment)
        cell.font = copy(pz_cell.font)
        cell.border = copy(pz_cell.border)
        cell.fill = copy(pz_cell.fill)

    def get_row_dimensions(self, row_num) -> RowDimension:
        dimension = self.sheet.row_dimensions[row_num]
        return dimension

    def set_row_dimensions(self, row_num, dimension: RowDimension):
        # self.sheet.row_dimensions[row_num].height = value.height
        if dimension.customHeight:
            self.sheet.row_dimensions[row_num].height = dimension.height
            # logger.trace(f"row {row_num} height: {dimension.height}, {self.sheet.row_dimensions[row_num].customHeight}")
        # self.sheet.row_dimensions[row_num].alignment = value.alignment

    def get_column_dimension(self, col_num) -> ColumnDimension:
        column_a_width = self.sheet.column_dimensions[col_num].width
        if column_a_width is None:
            logger.debug(f"Column {col_num} width is not set explicitly.")
        else:
            logger.debug(f"Column {col_num} width: {column_a_width}")
        logger.debug(f'get column dimension for {col_num}, width: {self.sheet.column_dimensions[col_num].width}')
        return self.sheet.column_dimensions[col_num]

    def set_column_dimension(self, col_num, value: ColumnDimension):
        self.sheet.column_dimensions[col_num].width = value.width
        pass

    def set_column_width(self, col_num, width: float):
        self.sheet.column_dimensions[get_column_letter(col_num)].width = width

    def get_headers(self) -> dict[str, int]:
        return self.headers

    def get_header_row(self) -> int:
        return self.header_row

    def max_column(self) -> int:
        return self.max_column

    def max_row(self) -> int:
        return self.sheet.max_row

    def save_as(self, filename: str):
        logger.info(f'Save excel workbook as: {filename}')
        self.wb.save(filename)

    def sheet_rename(self, new_name: str):
        self.sheet.title = new_name

    def duplicate_sheet(self, source: str, target: str):
        source_sheet = self.wb[source]
        new_sheet = self.wb.copy_worksheet(source_sheet)
        new_sheet.title = target
        # new_sheet.print_titles = source_sheet.print_titles
        # new_sheet.print_titles(source_sheet.print_titles)
        new_sheet.print_title_rows = source_sheet.print_title_rows
        new_sheet.print_title_cols = source_sheet.print_title_cols

        self.sheet = new_sheet
